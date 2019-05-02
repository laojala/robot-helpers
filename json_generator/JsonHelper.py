from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import json

class JsonHelper:

    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    __version__ = '0.1'

    def add_value_to_custom_dictionary(self, dictionary, key, value):
        if key not in dictionary:
            dictionary[key] = value
        elif type(dictionary[key]) == list:
            dictionary[key].append(value)
        else:
            dictionary[key] = [dictionary[key], value]

    def get_column_as_list(self, file, sheet, column_index):
        wb = load_workbook(filename = file)
        ws = wb[sheet]
        mylist = []
        for col in ws[get_column_letter(int(column_index)+1)]:
            if (col.value == None):
                mylist.append(u"")
            else:
                mylist.append(col.value)
        return mylist

    def get_column_count_custom(self, file, sheet):
        wb = load_workbook(filename = file)
        ws = wb[sheet]
        return ws.max_column

    def get_json_string(self, dict):
        return json.dumps(dict, indent=4, ensure_ascii=False).encode('utf8')
